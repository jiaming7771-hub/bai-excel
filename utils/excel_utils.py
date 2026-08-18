from __future__ import annotations

from datetime import date, datetime, time
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from utils.error_handler import AppError, friendly_error

# 这些列按文本写入，避免手机号/证件号变成科学计数法
TEXT_COLUMN_HINTS = (
    "手机",
    "电话",
    "联系方式",
    "mobile",
    "phone",
    "tel",
    "身份证",
    "证件号",
    "银行卡",
    "卡号",
    "订单号",
    "单号",
    "客户编号",
    "工号",
    "快递单",
)


def read_excel(path: str | Path) -> pd.DataFrame:
    file_path = Path(path)
    suffix = file_path.suffix.lower()
    try:
        if suffix == ".xls":
            df = pd.read_excel(file_path, engine="xlrd", dtype=object)
        else:
            df = pd.read_excel(file_path, engine="openpyxl", dtype=object)
    except Exception as exc:
        raise friendly_error(exc) from exc

    if df is None or df.empty:
        raise AppError("Excel 文件是空的", "请确认表格里至少有一行数据。")
    df.columns = [str(c).strip() for c in df.columns]
    if any(col.startswith("Unnamed") for col in df.columns) and len(df.columns) == 1:
        raise AppError("没有识别到表头", "请确认第一行是表头，并且表格内容完整。")
    return normalize_text_columns(df)


def write_excel(df: pd.DataFrame, path: str | Path, sheet_name: str = "Sheet1") -> Path:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = (sheet_name or "Sheet1")[:31]
        fill_worksheet(ws, df)
        wb.save(output)
    except PermissionError as exc:
        raise AppError(
            "无法保存结果文件",
            "请确认输出文件没有被 Excel 打开，并且您有权限写入这个文件夹。",
        ) from exc
    except Exception as exc:
        raise friendly_error(exc) from exc
    return output


def fill_worksheet(ws, df: pd.DataFrame) -> None:
    """把一张表写进工作表：文本列防变形 + 日期不带时分秒 + 表头加粗冻结 + 自动列宽。"""
    prepared = normalize_text_columns(df.copy())
    for row in dataframe_to_rows(prepared, index=False, header=True):
        ws.append([excel_value(value) for value in row])
    apply_text_column_format(ws, list(prepared.columns))
    apply_date_column_format(ws, list(prepared.columns))
    style_header_row(ws)
    autofit_worksheet(ws)
    ws.freeze_panes = "A2"


def list_columns(df: pd.DataFrame) -> list[str]:
    return [str(c) for c in df.columns]


def autofit_worksheet(ws, min_width: float = 10, max_width: float = 48, padding: float = 2.5) -> None:
    for index, column_cells in enumerate(ws.columns, start=1):
        lengths = [display_width(cell.value) for cell in column_cells if cell.value is not None]
        width = max(lengths) + padding if lengths else min_width
        width = max(min_width, min(width, max_width))
        ws.column_dimensions[get_column_letter(index)].width = width


def autofit_workbook(path: str | Path) -> None:
    workbook = load_workbook(path)
    for sheet in workbook.worksheets:
        autofit_worksheet(sheet)
        if sheet.max_row >= 1:
            style_header_row(sheet)
            sheet.freeze_panes = "A2"
    workbook.save(path)


def is_text_column(name: object) -> bool:
    text = str(name).strip().lower()
    return any(hint in text for hint in TEXT_COLUMN_HINTS)


def is_phone_column(name: object) -> bool:
    text = str(name).strip().lower()
    return any(hint in text for hint in ("手机", "电话", "联系方式", "mobile", "phone", "tel"))


def format_text_value(value) -> object:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass

    if isinstance(value, bool):
        return value
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    if isinstance(value, int):
        return str(value)

    text = str(value).strip()
    if not text or text.lower() in {"nan", "none", "nat", "<na>"}:
        return None
    if "e+" in text.lower() or "e-" in text.lower():
        try:
            number = float(text)
            if number.is_integer():
                return str(int(number))
        except ValueError:
            pass
    if text.endswith(".0") and text[:-2].replace("-", "", 1).isdigit():
        return text[:-2]
    return text


def format_phone_value(value) -> object:
    return format_text_value(value)


def normalize_text_columns(df: pd.DataFrame) -> pd.DataFrame:
    for column in df.columns:
        if is_text_column(column):
            df[column] = df[column].map(format_text_value)
    return df


def apply_text_column_format(ws, columns: list[str]) -> None:
    for index, column in enumerate(columns, start=1):
        if not is_text_column(column):
            continue
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=index)
            if cell.value is None or cell.value == "":
                continue
            cell.value = str(format_text_value(cell.value) or "")
            cell.number_format = "@"


def apply_date_column_format(ws, columns: list[str]) -> None:
    for index, _column in enumerate(columns, start=1):
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=index)
            value = cell.value
            if value is None or value == "":
                continue
            # 已经是 YYYY-MM-DD 文本
            if isinstance(value, str) and _looks_like_date_only(value):
                cell.number_format = "@"
                continue
            normalized = excel_value(value)
            if isinstance(normalized, str) and _looks_like_date_only(normalized):
                cell.value = normalized
                cell.number_format = "@"
            elif isinstance(normalized, datetime):
                cell.value = normalized
                cell.number_format = "YYYY-MM-DD HH:MM:SS"


def _looks_like_date_only(text: str) -> bool:
    value = text.strip()
    if len(value) != 10:
        return False
    try:
        datetime.strptime(value, "%Y-%m-%d")
        return True
    except ValueError:
        try:
            datetime.strptime(value, "%Y/%m/%d")
            return True
        except ValueError:
            return False


def style_header_row(ws) -> None:
    if ws.max_row < 1:
        return
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")


def _is_midnight(value: datetime) -> bool:
    return value.hour == 0 and value.minute == 0 and value.second == 0 and value.microsecond == 0


def excel_value(value):
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass

    # pandas Timestamp / numpy datetime64 → Python date/datetime
    if isinstance(value, pd.Timestamp):
        if pd.isna(value):
            return None
        if _is_midnight(value.to_pydatetime()):
            return value.date().strftime("%Y-%m-%d")
        return value.to_pydatetime()

    if isinstance(value, datetime):
        if _is_midnight(value):
            return value.strftime("%Y-%m-%d")
        return value

    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")

    if isinstance(value, time):
        return value

    # 字符串日期：只保留年月日，不带 0:00:00
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
            try:
                parsed = datetime.strptime(text, fmt)
                if _is_midnight(parsed) or " " not in fmt:
                    return parsed.strftime("%Y-%m-%d")
                return parsed
            except ValueError:
                continue
        # 常见显示：2024-01-15 0:00:00
        if " 0:00:00" in text or " 00:00:00" in text:
            head = text.split(" ")[0]
            for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
                try:
                    return datetime.strptime(head, fmt).strftime("%Y-%m-%d")
                except ValueError:
                    continue

    if hasattr(value, "item") and not isinstance(value, (bytes, str)):
        try:
            return excel_value(value.item())
        except Exception:
            return value
    return value


def display_width(value) -> float:
    text = str(value)
    width = 0.0
    for char in text:
        width += 2.0 if ord(char) > 127 else 1.0
    return width
