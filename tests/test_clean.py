from __future__ import annotations

import pandas as pd
from openpyxl import load_workbook

from core.excel_clean import clean_excel


def test_clean_removes_duplicates_and_blanks(testdata, tmp_path, monkeypatch):
    import app_config

    monkeypatch.setattr(app_config, "OUTPUT_DIR", tmp_path)
    original = pd.read_excel(testdata["clean"])
    result = clean_excel(
        testdata["clean"],
        drop_duplicates=True,
        drop_blank_rows=True,
        dedupe_column="手机号",
    )
    cleaned = pd.read_excel(result.output_path, sheet_name="清洗结果")
    deleted = pd.read_excel(result.report_path, sheet_name="已删除记录")
    assert result.original_rows == len(original)
    assert result.cleaned_rows == len(cleaned)
    assert result.deleted_rows == result.original_rows - result.cleaned_rows
    assert result.deleted_rows > 0
    assert len(deleted) == result.deleted_rows
    assert "删除原因" in deleted.columns
    assert cleaned["手机号"].dropna().duplicated().sum() == 0
    assert "清洗标记" not in cleaned.columns


def test_clean_fixes_and_flags_instead_of_deleting(tmp_path, monkeypatch):
    import app_config

    monkeypatch.setattr(app_config, "OUTPUT_DIR", tmp_path)
    source = tmp_path / "dirty.xlsx"
    pd.DataFrame(
        [
            {"姓名": "张三", "手机号": "13800138001", "邮箱": "a@test.com", "最近下单": "2024-01-02"},
            {"姓名": "13900139001", "手机号": "b@test.com", "邮箱": "李四", "最近下单": "2024/3/5"},  # 窜行可归位
            {"姓名": "王五", "手机号": "138-0013-8002", "邮箱": "C@TEST.COM", "最近下单": "20240308"},  # 可修复
            {"姓名": "赵六", "手机号": "12345", "邮箱": "not-email", "最近下单": "昨天"},  # 不可修复
            {"姓名": "abc", "手机号": "王七", "邮箱": "13800138888", "最近下单": "2024-01-01"},  # 窜行但夹杂无法识别，待核对
        ]
    ).to_excel(source, index=False)

    result = clean_excel(
        source,
        drop_duplicates=False,
        drop_blank_rows=False,
        check_shifted_rows=True,
        fix_phone=True,
        fix_email=True,
        fix_dates=True,
    )
    cleaned = pd.read_excel(result.output_path, sheet_name="清洗结果")
    review = pd.read_excel(result.report_path, sheet_name="待人工核对")
    fixed_log = pd.read_excel(result.report_path, sheet_name="已自动修复")
    quality = pd.read_excel(result.report_path, sheet_name="清洗报告")
    deleted = pd.read_excel(result.report_path, sheet_name="已删除记录")
    wb = load_workbook(result.report_path)
    ws = wb["核对视图"]

    assert "清洗报告" in wb.sheetnames
    assert "数据健康度" in quality["项目"].astype(str).tolist()
    assert result.quality_score <= 100
    assert result.quality_summary

    # 有问题的行不该被删光，应留在结果里
    assert len(cleaned) == 5
    assert len(deleted) == 0
    assert result.review_rows >= 1
    assert result.fixed_cells >= 3
    assert "问题说明" in review.columns
    assert list(fixed_log.columns) == ["结果行号", "字段", "原值", "新值", "修复说明"]
    assert len(fixed_log) >= 3
    assert any("窜行" in str(x) for x in fixed_log["修复说明"].tolist())
    # 窜行已归位
    row = cleaned[cleaned["姓名"] == "李四"].iloc[0]
    assert str(row["手机号"]) == "13900139001"
    assert "b@test.com" in str(row["邮箱"]).lower()
    assert "13800138002" in cleaned["手机号"].astype(str).tolist()
    assert "c@test.com" in cleaned["邮箱"].astype(str).tolist()
    assert "2024-03-08" in cleaned["最近下单"].astype(str).tolist()
    # 夹杂无法识别的窜行仍待核对
    assert any("窜行" in str(x) for x in review["问题说明"].tolist())

    fills = []
    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row_cells:
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                fills.append(str(cell.fill.fgColor.rgb))
    assert any("FFFF00" in f for f in fills)
    assert any("FF9800" in f for f in fills)
    assert "清洗标记" not in cleaned.columns
    review_view = pd.read_excel(result.report_path, sheet_name="核对视图")
    assert "清洗标记" in review_view.columns
    assert "已修复" in review_view["清洗标记"].astype(str).tolist()
    assert any("待核对" in str(x) for x in review_view["清洗标记"].tolist())
