from __future__ import annotations

import pandas as pd
from openpyxl import load_workbook

from core.sales_report import auto_detect_columns, build_sales_report


def test_sales_report_sheets_and_ranking(testdata, tmp_path, monkeypatch):
    import app_config

    monkeypatch.setattr(app_config, "OUTPUT_DIR", tmp_path)
    source = pd.read_excel(testdata["sales"])
    mapping = auto_detect_columns(list(source.columns))
    assert mapping["person"] == "销售人员"
    assert mapping["amount"] == "销售额"
    result = build_sales_report(
        testdata["sales"],
        person_col="销售人员",
        amount_col="销售额",
        qty_col="数量",
        product_col="产品",
        date_col="日期",
    )
    wb = load_workbook(result.output_path)
    assert "销售人员汇总" in wb.sheetnames
    assert "产品汇总" in wb.sheetnames
    assert "原始数据" in wb.sheetnames
    assert "月度销售" in wb.sheetnames

    person = pd.read_excel(result.output_path, sheet_name="销售人员汇总")
    assert list(person.columns) == ["销售人员", "销售额", "销量", "排名"]
    assert person["排名"].tolist() == list(range(1, len(person) + 1))
    assert person["销售额"].is_monotonic_decreasing
    assert result.people_count == 20
    assert result.product_count == 30

    monthly = pd.read_excel(result.output_path, sheet_name="月度销售")
    assert len(monthly) == 12
    ws = wb["月度销售"]
    assert len(ws._charts) == 1
