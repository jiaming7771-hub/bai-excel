from __future__ import annotations

from openpyxl import load_workbook

from core.excel_clean import clean_excel
from core.excel_merge import merge_excels
from core.excel_split import split_excel
from core.sales_report import build_sales_report
from utils.excel_utils import write_excel


def _assert_sheet_formatted(ws) -> None:
    assert ws.freeze_panes == "A2"
    assert ws["A1"].font.bold is True
    assert ws.column_dimensions["A"].width >= 10


def test_write_excel_autofits_and_keeps_phone_text(tmp_path):
    import pandas as pd

    path = tmp_path / "demo.xlsx"
    df = pd.DataFrame(
        [
            {"姓名": "张三", "手机号": 13800138001, "销售额": 12800},
            {"姓名": "钱十一", "手机号": 13900139002, "销售额": 21000},
        ]
    )
    write_excel(df, path)
    wb = load_workbook(path)
    ws = wb.active
    _assert_sheet_formatted(ws)
    assert ws.column_dimensions["B"].width >= 12
    assert ws["B2"].value == "13800138001"
    assert ws["B2"].number_format == "@"
    assert "E+" not in str(ws["B2"].value)


def test_write_excel_keeps_date_without_time(tmp_path):
    from datetime import date, datetime

    import pandas as pd

    path = tmp_path / "dates.xlsx"
    df = pd.DataFrame(
        [
            {"姓名": "周杰", "入职日期": datetime(2024, 1, 15)},
            {"姓名": "马芳", "入职日期": date(2023, 9, 9)},
            {"姓名": "罗斌", "入职日期": "2021-12-03 0:00:00"},
        ]
    )
    write_excel(df, path)
    wb = load_workbook(path)
    ws = wb.active
    assert ws["B2"].value == "2024-01-15"
    assert ws["B3"].value == "2023-09-09"
    assert ws["B4"].value == "2021-12-03"
    assert "0:00:00" not in str(ws["B2"].value)
    assert "0:00:00" not in str(ws["B4"].value)


def test_split_output_dates_without_time(testdata, tmp_path, monkeypatch):
    from datetime import date, datetime

    import app_config
    import pandas as pd

    monkeypatch.setattr(app_config, "OUTPUT_DIR", tmp_path)
    source = tmp_path / "staff.xlsx"
    pd.DataFrame(
        [
            {"姓名": "周杰", "部门": "华北", "入职日期": date(2024, 1, 15)},
            {"姓名": "马芳", "部门": "华北", "入职日期": date(2023, 9, 9)},
            {"姓名": "陈晨", "部门": "华东", "入职日期": datetime(2023, 3, 1)},
        ]
    ).to_excel(source, index=False)
    result = split_excel(source, "部门")
    north = load_workbook(result.output_dir / "华北.xlsx").active
    assert north["C2"].value == "2024-01-15"
    assert "0:00:00" not in str(north["C2"].value)
    import app_config

    monkeypatch.setattr(app_config, "OUTPUT_DIR", tmp_path)

    merged = merge_excels([testdata["sales_a"], testdata["sales_b"], testdata["sales_c"]])
    split = split_excel(testdata["split"], "部门")
    cleaned = clean_excel(
        testdata["clean"],
        drop_duplicates=True,
        drop_blank_rows=True,
        dedupe_column="手机号",
    )
    sales = build_sales_report(
        testdata["sales"],
        person_col="销售人员",
        amount_col="销售额",
        qty_col="数量",
        product_col="产品",
        date_col="日期",
    )

    merge_ws = load_workbook(merged.output_path).active
    clean_ws = load_workbook(cleaned.output_path).active
    sales_wb = load_workbook(sales.output_path)
    split_files = list(split.output_dir.glob("*.xlsx"))

    _assert_sheet_formatted(merge_ws)
    _assert_sheet_formatted(clean_ws)
    assert split_files
    _assert_sheet_formatted(load_workbook(split_files[0]).active)
    for name in ("销售人员汇总", "产品汇总", "原始数据", "月度销售"):
        _assert_sheet_formatted(sales_wb[name])
