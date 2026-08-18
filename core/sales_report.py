from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook

from app_config import ensure_output_dir
from utils.excel_utils import read_excel
from utils.error_handler import AppError


PERSON_ALIASES = ["销售人员", "业务员", "销售", "姓名", "员工", "负责人"]
AMOUNT_ALIASES = ["销售额", "金额", "成交额", "营收", "销售金额"]
QTY_ALIASES = ["数量", "销量", "件数"]
PRODUCT_ALIASES = ["产品", "商品", "品名", "SKU", "货品"]
DATE_ALIASES = ["日期", "成交日期", "下单日期", "销售日期", "时间"]


@dataclass
class SalesResult:
    output_path: Path
    people_count: int
    product_count: int
    row_count: int


def auto_detect_columns(columns: list[str]) -> dict[str, str | None]:
    return {
        "person": _match(columns, PERSON_ALIASES),
        "amount": _match(columns, AMOUNT_ALIASES),
        "qty": _match(columns, QTY_ALIASES),
        "product": _match(columns, PRODUCT_ALIASES),
        "date": _match(columns, DATE_ALIASES),
    }


def _match(columns: list[str], aliases: list[str]) -> str | None:
    for alias in aliases:
        for col in columns:
            if str(col).strip() == alias:
                return col
    for alias in aliases:
        for col in columns:
            if alias in str(col):
                return col
    return None


def _to_number(series: pd.Series) -> pd.Series:
    cleaned = (
        series.astype(str)
        .str.replace(",", "", regex=False)
        .str.replace("￥", "", regex=False)
        .str.replace("¥", "", regex=False)
        .str.replace("元", "", regex=False)
        .str.strip()
    )
    return pd.to_numeric(cleaned, errors="coerce").fillna(0)


def build_sales_report(
    path: Path,
    person_col: str,
    amount_col: str,
    qty_col: str | None = None,
    product_col: str | None = None,
    date_col: str | None = None,
    output_name: str = "sales_report.xlsx",
    progress_cb=None,
) -> SalesResult:
    if progress_cb:
        progress_cb(8, "正在读取 Excel")
    df = read_excel(path)
    for required, label in ((person_col, "销售人员"), (amount_col, "销售额")):
        if not required or required not in df.columns:
            raise AppError(f"请选择{label}字段", "字段名称如果不一样，请手动选择正确的列。")

    work = df.copy()
    work["_amount"] = _to_number(work[amount_col])
    work["_qty"] = _to_number(work[qty_col]) if qty_col and qty_col in work.columns else 1

    if progress_cb:
        progress_cb(35, "正在汇总销售人员")
    person = (
        work.groupby(person_col, dropna=False, as_index=False)
        .agg(销售额=("_amount", "sum"), 销量=("_qty", "sum"))
        .sort_values("销售额", ascending=False)
    )
    person["排名"] = range(1, len(person) + 1)
    person = person.rename(columns={person_col: "销售人员"})[["销售人员", "销售额", "销量", "排名"]]

    if progress_cb:
        progress_cb(55, "正在汇总产品")
    if product_col and product_col in work.columns:
        product = (
            work.groupby(product_col, dropna=False, as_index=False)
            .agg(销售额=("_amount", "sum"), 销量=("_qty", "sum"))
            .sort_values("销售额", ascending=False)
            .rename(columns={product_col: "产品"})
        )
    else:
        product = pd.DataFrame(columns=["产品", "销售额", "销量"])

    monthly = pd.DataFrame(columns=["月份", "销售额", "销量"])
    if date_col and date_col in work.columns:
        if progress_cb:
            progress_cb(70, "正在汇总月度销售")
        dates = pd.to_datetime(work[date_col], errors="coerce")
        work["_month"] = dates.dt.to_period("M").astype(str)
        monthly = (
            work.dropna(subset=["_month"])
            .groupby("_month", as_index=False)
            .agg(销售额=("_amount", "sum"), 销量=("_qty", "sum"))
            .rename(columns={"_month": "月份"})
            .sort_values("月份")
        )

    output = ensure_output_dir() / output_name
    if progress_cb:
        progress_cb(88, "正在生成报表")
    _write_workbook(output, person, product, df, monthly)
    if progress_cb:
        progress_cb(100, "处理完成！")
    return SalesResult(
        output_path=output,
        people_count=len(person),
        product_count=len(product),
        row_count=len(df),
    )


def _write_workbook(
    path: Path,
    person: pd.DataFrame,
    product: pd.DataFrame,
    original: pd.DataFrame,
    monthly: pd.DataFrame,
) -> None:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "销售人员汇总"
    _fill_sheet(ws1, person)

    ws2 = wb.create_sheet("产品汇总")
    _fill_sheet(ws2, product)

    ws3 = wb.create_sheet("原始数据")
    _fill_sheet(ws3, original)

    if not monthly.empty:
        ws4 = wb.create_sheet("月度销售")
        _fill_sheet(ws4, monthly)
        chart = BarChart()
        chart.type = "col"
        chart.title = "月度销售额"
        chart.y_axis.title = "销售额"
        chart.x_axis.title = "月份"
        data = Reference(ws4, min_col=2, min_row=1, max_row=ws4.max_row, max_col=2)
        cats = Reference(ws4, min_col=1, min_row=2, max_row=ws4.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        chart.width = 18
        chart.height = 8
        ws4.add_chart(chart, "E2")

    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(path)
    except PermissionError as exc:
        raise AppError(
            "无法保存结果文件",
            "请确认输出文件没有被 Excel 打开，并且您有权限写入这个文件夹。",
        ) from exc


def _fill_sheet(ws, df: pd.DataFrame) -> None:
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(list(row))
