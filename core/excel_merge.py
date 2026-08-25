from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from app_config import ensure_output_dir
from utils.excel_utils import fill_worksheet, normalize_header_name, read_excel, write_excel
from utils.error_handler import AppError
from utils.output_paths import merge_paths

# 同义词映射：把常见别名对齐到统一列名，减少「手机 vs 手机号」对不齐
HEADER_ALIASES = {
    "手机": "手机号",
    "电话": "手机号",
    "联系电话": "手机号",
    "mobile": "手机号",
    "phone": "手机号",
    "tel": "手机号",
    "客户名": "姓名",
    "联系人": "姓名",
    "名字": "姓名",
    "邮件": "邮箱",
    "email": "邮箱",
    "mail": "邮箱",
    "e-mail": "邮箱",
    "销售金额": "销售额",
    "金额": "销售额",
    "营收": "销售额",
}

ORANGE = PatternFill(fill_type="solid", fgColor="FFFF9800")
YELLOW = PatternFill(fill_type="solid", fgColor="FFFFFF00")


@dataclass
class MergeResult:
    output_path: Path
    report_path: Path
    file_count: int
    row_count: int
    warning_count: int = 0
    detail_text: str = ""


@dataclass
class _FileMeta:
    name: str
    rows: int
    orig_to_new: dict[str, str]


def _align_headers(df: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, str]]:
    renamed: dict[str, str] = {}
    used: set[str] = set()
    orig_to_new: dict[str, str] = {}
    for col in df.columns:
        raw = normalize_header_name(col)
        mapped = HEADER_ALIASES.get(raw.lower(), HEADER_ALIASES.get(raw, raw))
        name = mapped
        if name in used:
            index = 2
            while f"{name}_{index}" in used:
                index += 1
            name = f"{name}_{index}"
        used.add(name)
        renamed[col] = name
        orig_to_new[str(col)] = name
    out = df.rename(columns=renamed)
    out.columns = [str(c).strip() for c in out.columns]
    return out, orig_to_new


def _build_alignment_warnings(
    frames: list[pd.DataFrame],
    metas: list[_FileMeta],
    columns: list[str],
    source_col: str,
) -> list[str]:
    warnings: list[str] = []
    data_cols = [c for c in columns if c != source_col]
    if len(metas) <= 1:
        return warnings

    filled_map: dict[str, set[str]] = {c: set() for c in data_cols}
    for meta, frame in zip(metas, frames):
        for col in data_cols:
            if col not in frame.columns:
                continue
            series = frame[col]
            if series.isna().all():
                continue
            text = series.astype(str).str.strip()
            if (series.notna() & (text != "") & (text.str.lower() != "nan")).any():
                filled_map[col].add(meta.name)

    all_files = {m.name for m in metas}
    for col, files in filled_map.items():
        if files and len(files) < len(all_files):
            missing = "、".join(sorted(all_files - files))
            warnings.append(f"「{col}」仅在部分文件有数据，以下文件该列为空：{missing}")

    for i, c1 in enumerate(data_cols):
        f1 = filled_map.get(c1, set())
        if not f1:
            continue
        for c2 in data_cols[i + 1 :]:
            f2 = filled_map.get(c2, set())
            if not f2:
                continue
            if f1.isdisjoint(f2) and f1 | f2 == all_files:
                warnings.append(
                    f"疑似未对齐：「{c1}」与「{c2}」各在不同文件有数据，可能是同一字段但列名不同"
                )
    return warnings


def _build_report_frames(
    metas: list[_FileMeta],
    columns: list[str],
    warnings: list[str],
    source_col: str,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    summary_rows = [
        {"项目": "合并文件数", "内容": len(metas)},
        {"项目": "合并后总行数", "内容": sum(m.rows for m in metas)},
        {"项目": "合并后总列数", "内容": len(columns)},
        {"项目": "发现问题", "内容": len(warnings)},
        {
            "项目": "怎么看",
            "内容": "先看本页警告 → 列名对照表看各文件列名 → 空值统计看哪列在哪张表是空的",
        },
    ]
    for idx, msg in enumerate(warnings, start=1):
        summary_rows.append({"项目": f"警告{idx}", "内容": msg})
    summary = pd.DataFrame(summary_rows)

    matrix_rows: list[dict] = []
    for col in columns:
        if col == source_col:
            continue
        row: dict = {"合并后列名": col}
        for meta in metas:
            originals = [o for o, n in meta.orig_to_new.items() if n == col]
            if originals:
                labels = []
                for o in originals:
                    labels.append(o if o == col else f"{o}→{col}")
                row[meta.name] = "；".join(labels)
            else:
                row[meta.name] = "—（本表无此列）"
        matrix_rows.append(row)
    matrix = pd.DataFrame(matrix_rows)
    return summary, matrix


def _build_null_stats(
    frames: list[pd.DataFrame],
    metas: list[_FileMeta],
    columns: list[str],
    source_col: str,
) -> pd.DataFrame:
    rows: list[dict] = []
    for meta, frame in zip(metas, frames):
        for col in columns:
            if col == source_col:
                continue
            if col not in frame.columns:
                rows.append(
                    {
                        "文件": meta.name,
                        "列名": col,
                        "本文件行数": meta.rows,
                        "有数据行数": 0,
                        "空值行数": meta.rows,
                        "说明": "本表无此列",
                    }
                )
                continue
            series = frame[col]
            text = series.astype(str).str.strip()
            has = (series.notna() & (text != "") & (text.str.lower() != "nan")).sum()
            empty = meta.rows - int(has)
            note = "正常" if empty == 0 else ("全空" if has == 0 else "部分为空")
            rows.append(
                {
                    "文件": meta.name,
                    "列名": col,
                    "本文件行数": meta.rows,
                    "有数据行数": int(has),
                    "空值行数": int(empty),
                    "说明": note,
                }
            )
    return pd.DataFrame(rows)


def _write_merge_result(path: Path, merged: pd.DataFrame) -> None:
    write_excel(merged, path, sheet_name="合并结果")


def _write_merge_report(
    path: Path,
    summary: pd.DataFrame,
    matrix: pd.DataFrame,
    null_stats: pd.DataFrame,
    warnings: list[str],
) -> None:
    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "列对齐报告"
    fill_worksheet(ws_sum, summary)
    warn_start = 6
    for row_idx in range(warn_start, ws_sum.max_row + 1):
        label = ws_sum.cell(row=row_idx, column=1).value
        if label and str(label).startswith("警告"):
            for col_idx in range(1, ws_sum.max_column + 1):
                ws_sum.cell(row=row_idx, column=col_idx).fill = ORANGE

    ws_map = wb.create_sheet("列名对照")
    fill_worksheet(ws_map, matrix)

    ws_null = wb.create_sheet("空值统计")
    fill_worksheet(ws_null, null_stats)
    if not null_stats.empty:
        note_col = list(null_stats.columns).index("说明") + 1
        for row_idx in range(2, ws_null.max_row + 1):
            note = ws_null.cell(row=row_idx, column=note_col).value
            if note in {"本表无此列", "全空", "部分为空"}:
                for col_idx in range(1, ws_null.max_column + 1):
                    ws_null.cell(row=row_idx, column=col_idx).fill = YELLOW

    ws_help = wb.create_sheet("说明")
    help_df = pd.DataFrame(
        [
            {"项目": "本文件用途", "内容": "内部核对列对齐与空值，勿直接外传"},
            {"项目": "结果文件", "内容": "同目录下的「合并结果_*.xlsx」仅含合并数据，可发给同事"},
            {"项目": "发现问题", "内容": len(warnings)},
            {"项目": "建议", "内容": "有橙色警告时，按「列名对照」「空值统计」回到源表改列名或补数据"},
        ]
    )
    fill_worksheet(ws_help, help_df)

    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(path)
    except PermissionError as exc:
        raise AppError(
            "无法保存处理报告",
            "请确认报告文件没有被 Excel 打开，并且您有权限写入这个文件夹。",
        ) from exc


def merge_excels(
    paths: list[Path],
    output_name: str | None = None,
    add_source_column: bool = True,
    progress_cb=None,
) -> MergeResult:
    if not paths:
        raise AppError("请先选择 Excel 文件", "可以一次选择多个文件，也可以选择一个包含 Excel 的文件夹。")

    frames: list[pd.DataFrame] = []
    metas: list[_FileMeta] = []
    columns: list[str] = []
    seen: set[str] = set()
    used_files = 0
    total = max(len(paths), 1)
    source_col = "来源文件"

    for index, path in enumerate(paths, start=1):
        if progress_cb:
            progress_cb(int(index / (total + 1) * 80), f"正在读取 {path.name}")
        try:
            raw_df = read_excel(path)
            df, orig_to_new = _align_headers(raw_df)
        except AppError as exc:
            if "空" in exc.title:
                continue
            raise
        used_files += 1
        if add_source_column:
            if source_col in df.columns:
                df = df.rename(columns={source_col: f"{source_col}_原表"})
            df.insert(0, source_col, path.name)
        frames.append(df)
        metas.append(_FileMeta(name=path.name, rows=len(df), orig_to_new=orig_to_new))
        for col in df.columns:
            if col not in seen:
                seen.add(col)
                columns.append(col)

    if not frames:
        raise AppError("这些文件都没有可合并的数据", "请确认每个 Excel 至少有一行表头和数据。")

    if progress_cb:
        progress_cb(88, "正在按列名对齐并合并")

    if add_source_column and source_col in columns:
        columns = [source_col] + [c for c in columns if c != source_col]

    warnings = _build_alignment_warnings(frames, metas, columns, source_col)
    summary, matrix = _build_report_frames(metas, columns, warnings, source_col)
    null_stats = _build_null_stats(frames, metas, columns, source_col)

    aligned = [frame.reindex(columns=columns) for frame in frames]
    merged = pd.concat(aligned, ignore_index=True, sort=False)

    stem = Path(paths[0]).stem if len(paths) == 1 else "多文件合并"
    result_path, report_path = merge_paths(stem)
    if output_name:
        result_path = ensure_output_dir() / output_name

    if progress_cb:
        progress_cb(95, "正在保存结果与处理报告")
    _write_merge_result(result_path, merged)
    _write_merge_report(report_path, summary, matrix, null_stats, warnings)
    if progress_cb:
        progress_cb(100, "处理完成！")

    if warnings:
        detail = (
            f"结果 {len(merged)} 行已保存，可外传；"
            f"发现 {len(warnings)} 个列对齐问题，请打开「处理报告」核对"
        )
    else:
        detail = f"结果 {len(merged)} 行已保存；列名均已对齐，详见「处理报告」"

    return MergeResult(
        output_path=result_path,
        report_path=report_path,
        file_count=used_files,
        row_count=len(merged),
        warning_count=len(warnings),
        detail_text=detail,
    )
