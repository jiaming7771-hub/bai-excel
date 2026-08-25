from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from utils.excel_utils import fill_worksheet, list_sheet_names, read_excel, write_excel
from utils.error_handler import AppError
from utils.file_utils import sanitize_filename
from utils.output_paths import split_paths

ORANGE = PatternFill(fill_type="solid", fgColor="FFFF9800")
YELLOW = PatternFill(fill_type="solid", fgColor="FFFFFF00")


@dataclass
class SplitFileInfo:
    filename: str
    label: str
    rows: int
    note: str = ""


@dataclass
class SplitResult:
    output_dir: Path
    report_path: Path
    file_count: int
    row_count: int
    mode: str = "column"
    warning_count: int = 0
    detail_text: str = ""
    file_infos: list[SplitFileInfo] = field(default_factory=list)


def _unique_filename(base: str, used_names: dict[str, int]) -> tuple[str, bool]:
    safe = sanitize_filename(base, fallback="空值")
    count = used_names.get(safe, 0) + 1
    used_names[safe] = count
    duplicated = count > 1
    name = f"{safe}.xlsx" if count == 1 else f"{safe}_{count}.xlsx"
    return name, duplicated


def _blank_split_value(value) -> bool:
    if value is None:
        return True
    try:
        if pd.isna(value):
            return True
    except (TypeError, ValueError):
        pass
    return str(value).strip().lower() in {"", "nan", "none", "nat"}


def _write_split_report(
    path: Path,
    *,
    source_name: str,
    mode: str,
    mode_label: str,
    total_rows: int,
    file_infos: list[SplitFileInfo],
    skipped_sheets: list[str],
    manual_df: pd.DataFrame,
    warnings: list[str],
) -> None:
    summary_rows = [
        {"项目": "原文件", "内容": source_name},
        {"项目": "拆分方式", "内容": mode_label},
        {"项目": "源表总行数", "内容": total_rows},
        {"项目": "生成文件数", "内容": len(file_infos)},
        {"项目": "发现问题", "内容": len(warnings)},
        {
            "项目": "怎么看",
            "内容": "本报告用于核对；拆分出的 xlsx 在「拆分结果」文件夹，可直接使用",
        },
    ]
    for idx, msg in enumerate(warnings, start=1):
        summary_rows.append({"项目": f"警告{idx}", "内容": msg})
    if skipped_sheets:
        summary_rows.append(
            {"项目": "跳过的空表", "内容": "、".join(skipped_sheets)}
        )

    summary = pd.DataFrame(summary_rows)
    files_df = pd.DataFrame(
        [
            {
                "输出文件": info.filename,
                "分组/批次": info.label,
                "行数": info.rows,
                "备注": info.note,
            }
            for info in file_infos
        ]
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "拆分报告"
    fill_worksheet(ws, summary)
    for row_idx in range(2, ws.max_row + 1):
        label = ws.cell(row=row_idx, column=1).value
        if label and str(label).startswith("警告"):
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = ORANGE

    ws_files = wb.create_sheet("文件清单")
    fill_worksheet(ws_files, files_df)

    ws_manual = wb.create_sheet("待人工核对")
    fill_worksheet(ws_manual, manual_df)
    if not manual_df.empty:
        for row_idx in range(2, ws_manual.max_row + 1):
            for col_idx in range(1, ws_manual.max_column + 1):
                ws_manual.cell(row=row_idx, column=col_idx).fill = ORANGE

    ws_help = wb.create_sheet("说明")
    fill_worksheet(
        ws_help,
        pd.DataFrame(
            [
                {"项目": "结果文件夹", "内容": "仅含拆分后的数据文件，可外传"},
                {"项目": "本报告", "内容": "记录空值分组、跳过的工作表、重名文件等，供内部核对"},
            ]
        ),
    )

    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(path)
    except PermissionError as exc:
        raise AppError(
            "无法保存拆分报告",
            "请确认报告文件没有被 Excel 打开，并且您有权限写入这个文件夹。",
        ) from exc


def split_excel(
    path: Path,
    column: str | None = None,
    *,
    mode: str = "column",
    rows_per_file: int = 1000,
    progress_cb=None,
) -> SplitResult:
    """mode: column | sheet | rows"""
    mode = (mode or "column").strip().lower()
    if mode not in {"column", "sheet", "rows"}:
        raise AppError("不支持的拆分方式", "请选择：按字段、按工作表、或按行数。")

    result_dir, report_path = split_paths(path.stem)

    if mode == "sheet":
        return _split_by_sheets(path, result_dir, report_path, progress_cb)
    if mode == "rows":
        return _split_by_rows(path, result_dir, report_path, rows_per_file, progress_cb)
    if not column:
        raise AppError("请选择拆分字段", "例如：部门、城市、销售人员。")
    return _split_by_column(path, column, result_dir, report_path, progress_cb)


def _split_by_column(
    path: Path,
    column: str,
    output_dir: Path,
    report_path: Path,
    progress_cb,
) -> SplitResult:
    if progress_cb:
        progress_cb(5, "正在读取 Excel")
    df = read_excel(path)
    if column not in df.columns:
        raise AppError("找不到拆分字段", "请重新选择一个表格里实际存在的字段。")

    blank_mask = df[column].apply(_blank_split_value)
    manual_df = df[blank_mask].copy()
    work_df = df[~blank_mask].copy()

    groups = list(work_df.groupby(column, dropna=False, sort=False))
    if work_df.empty and manual_df.empty:
        raise AppError("没有可以拆分的数据", "请确认表格里至少有一行数据。")

    used_names: dict[str, int] = {}
    file_infos: list[SplitFileInfo] = []
    warnings: list[str] = []

    if not manual_df.empty:
        filename, dup = _unique_filename("空值", used_names)
        write_excel(manual_df.reset_index(drop=True), output_dir / filename)
        file_infos.append(
            SplitFileInfo(
                filename=filename,
                label="（拆分字段为空）",
                rows=len(manual_df),
                note="待人工补全拆分字段",
            )
        )
        warnings.append(f"有 {len(manual_df)} 行拆分字段为空，已写入「{filename}」，请人工核对")

    total = max(len(groups), 1)
    for index, (value, group) in enumerate(groups, start=1):
        label = str(value)
        filename, dup = _unique_filename(label, used_names)
        note = "文件名重复已自动加后缀" if dup else ""
        if dup:
            warnings.append(f"分组「{label}」输出文件名重复，已保存为 {filename}")
        write_excel(group.reset_index(drop=True), output_dir / filename)
        file_infos.append(SplitFileInfo(filename=filename, label=label, rows=len(group), note=note))
        if progress_cb:
            progress_cb(10 + int(index / total * 80), f"正在生成 {filename}")

    if progress_cb:
        progress_cb(92, "正在生成拆分报告")
    _write_split_report(
        report_path,
        source_name=path.name,
        mode="column",
        mode_label=f"按字段「{column}」",
        total_rows=len(df),
        file_infos=file_infos,
        skipped_sheets=[],
        manual_df=manual_df,
        warnings=warnings,
    )
    if progress_cb:
        progress_cb(100, "处理完成！")

    detail = f"已拆成 {len(file_infos)} 个文件，共 {len(df)} 行"
    if warnings:
        detail += f"；发现 {len(warnings)} 项需关注，请打开「处理报告」"

    return SplitResult(
        output_dir=output_dir,
        report_path=report_path,
        file_count=len(file_infos),
        row_count=len(df),
        mode="column",
        warning_count=len(warnings),
        detail_text=detail,
        file_infos=file_infos,
    )


def _split_by_sheets(
    path: Path,
    output_dir: Path,
    report_path: Path,
    progress_cb,
) -> SplitResult:
    sheets = list_sheet_names(path)
    if not sheets:
        raise AppError("没有找到工作表", "请确认 Excel 里至少有一个 Sheet。")

    used_names: dict[str, int] = {}
    file_infos: list[SplitFileInfo] = []
    skipped: list[str] = []
    warnings: list[str] = []
    total_rows = 0
    total = len(sheets)

    for index, sheet in enumerate(sheets, start=1):
        if progress_cb:
            progress_cb(5 + int(index / total * 80), f"正在拆分工作表：{sheet}")
        try:
            df = read_excel(path, sheet_name=sheet)
        except AppError as exc:
            if "空" in exc.title:
                skipped.append(sheet)
                continue
            raise
        filename, dup = _unique_filename(sheet, used_names)
        note = "文件名重复已自动加后缀" if dup else ""
        if dup:
            warnings.append(f"工作表「{sheet}」输出文件名重复，已保存为 {filename}")
        write_excel(df, output_dir / filename, sheet_name=sheet[:31])
        file_infos.append(SplitFileInfo(filename=filename, label=sheet, rows=len(df), note=note))
        total_rows += len(df)

    if not file_infos:
        raise AppError("没有可拆分的工作表数据", "请确认每个 Sheet 至少有一行表头和数据。")
    if skipped:
        warnings.append(f"已跳过 {len(skipped)} 个空工作表：{'、'.join(skipped)}")

    if progress_cb:
        progress_cb(92, "正在生成拆分报告")
    _write_split_report(
        report_path,
        source_name=path.name,
        mode="sheet",
        mode_label="按工作表",
        total_rows=total_rows,
        file_infos=file_infos,
        skipped_sheets=skipped,
        manual_df=pd.DataFrame(),
        warnings=warnings,
    )
    if progress_cb:
        progress_cb(100, "处理完成！")

    detail = f"已拆成 {len(file_infos)} 个文件，共 {total_rows} 行"
    if warnings:
        detail += f"；发现 {len(warnings)} 项需关注，请打开「处理报告」"

    return SplitResult(
        output_dir=output_dir,
        report_path=report_path,
        file_count=len(file_infos),
        row_count=total_rows,
        mode="sheet",
        warning_count=len(warnings),
        detail_text=detail,
        file_infos=file_infos,
    )


def _split_by_rows(
    path: Path,
    output_dir: Path,
    report_path: Path,
    rows_per_file: int,
    progress_cb,
) -> SplitResult:
    if rows_per_file < 1:
        raise AppError("每文件行数无效", "请输入大于 0 的整数，例如 1000。")
    if progress_cb:
        progress_cb(5, "正在读取 Excel")
    df = read_excel(path)
    if df.empty:
        raise AppError("没有可以拆分的数据", "请确认表格里至少有一行数据。")

    chunks = [df.iloc[i : i + rows_per_file] for i in range(0, len(df), rows_per_file)]
    used_names: dict[str, int] = {}
    file_infos: list[SplitFileInfo] = []
    warnings: list[str] = []
    total = len(chunks)

    for index, chunk in enumerate(chunks, start=1):
        label = f"第{index}批"
        filename, dup = _unique_filename(f"{label}_{rows_per_file}行", used_names)
        note = ""
        if len(chunk) < rows_per_file:
            note = f"本批仅 {len(chunk)} 行（未满 {rows_per_file} 行）"
            warnings.append(f"{filename}：{note}")
        write_excel(chunk.reset_index(drop=True), output_dir / filename)
        file_infos.append(SplitFileInfo(filename=filename, label=label, rows=len(chunk), note=note))
        if progress_cb:
            progress_cb(10 + int(index / total * 85), f"正在生成 {filename}")

    if progress_cb:
        progress_cb(92, "正在生成拆分报告")
    _write_split_report(
        report_path,
        source_name=path.name,
        mode="rows",
        mode_label=f"按行数（每文件 {rows_per_file} 行）",
        total_rows=len(df),
        file_infos=file_infos,
        skipped_sheets=[],
        manual_df=pd.DataFrame(),
        warnings=warnings,
    )
    if progress_cb:
        progress_cb(100, "处理完成！")

    detail = f"已拆成 {len(file_infos)} 个文件，共 {len(df)} 行"
    if warnings:
        detail += f"；末批可能不满额，详见「处理报告」"

    return SplitResult(
        output_dir=output_dir,
        report_path=report_path,
        file_count=total,
        row_count=len(df),
        mode="rows",
        warning_count=len(warnings),
        detail_text=detail,
        file_infos=file_infos,
    )
