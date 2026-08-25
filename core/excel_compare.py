from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from app_config import ensure_output_dir
from utils.excel_utils import fill_worksheet, read_excel, write_excel
from utils.error_handler import AppError
from utils.output_paths import compare_paths

YELLOW = PatternFill(fill_type="solid", fgColor="FFFFFF00")
ORANGE = PatternFill(fill_type="solid", fgColor="FFFF9800")
GREEN = PatternFill(fill_type="solid", fgColor="FFC8E6C9")


@dataclass
class CompareResult:
    output_path: Path
    report_path: Path
    only_a: int
    only_b: int
    changed: int
    same: int
    detail_text: str = ""


def _norm_key_part(value) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    return str(value).strip()


def _row_key(row: pd.Series, key_columns: list[str]) -> str:
    return "||".join(_norm_key_part(row.get(col)) for col in key_columns)


def _cell_text(value) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    return str(value).strip()


def compare_excels(
    path_a: Path,
    path_b: Path,
    key_columns: list[str],
    output_name: str | None = None,
    progress_cb=None,
) -> CompareResult:
    if not key_columns:
        raise AppError("请选择对比主键", "例如选择手机号、订单号，用来判断是不是同一条数据。")

    if progress_cb:
        progress_cb(8, "正在读取表 A")
    df_a = read_excel(path_a).copy()
    if progress_cb:
        progress_cb(20, "正在读取表 B")
    df_b = read_excel(path_b).copy()

    for col in key_columns:
        if col not in df_a.columns:
            raise AppError("表 A 找不到主键列", f"表 A 里没有「{col}」，请重新选择主键。")
        if col not in df_b.columns:
            raise AppError("表 B 找不到主键列", f"表 B 里没有「{col}」，请重新选择主键。")

    if progress_cb:
        progress_cb(40, "正在按主键匹配")

    compare_cols = sorted(set(df_a.columns) | set(df_b.columns))
    # 主键列靠前
    ordered_cols = list(key_columns) + [c for c in compare_cols if c not in key_columns]

    map_a: dict[str, list[int]] = {}
    map_b: dict[str, list[int]] = {}
    for idx, row in df_a.iterrows():
        map_a.setdefault(_row_key(row, key_columns), []).append(idx)
    for idx, row in df_b.iterrows():
        map_b.setdefault(_row_key(row, key_columns), []).append(idx)

    keys_a = set(map_a)
    keys_b = set(map_b)
    only_a_keys = sorted(keys_a - keys_b)
    only_b_keys = sorted(keys_b - keys_a)
    both_keys = sorted(keys_a & keys_b)

    only_a_rows: list[dict] = []
    only_b_rows: list[dict] = []
    changed_rows: list[dict] = []
    same_count = 0

    for key in only_a_keys:
        for idx in map_a[key]:
            row = {col: df_a.at[idx, col] if col in df_a.columns else "" for col in ordered_cols}
            row["对比结果"] = "仅在表A"
            only_a_rows.append(row)

    for key in only_b_keys:
        for idx in map_b[key]:
            row = {col: df_b.at[idx, col] if col in df_b.columns else "" for col in ordered_cols}
            row["对比结果"] = "仅在表B"
            only_b_rows.append(row)

    if progress_cb:
        progress_cb(70, "正在比对字段差异")

    for key in both_keys:
        # 同主键多行时，按出现顺序两两比对，多出来的进仅在A/仅在B
        idxs_a = map_a[key]
        idxs_b = map_b[key]
        pair_count = min(len(idxs_a), len(idxs_b))
        for i in range(pair_count):
            idx_a = idxs_a[i]
            idx_b = idxs_b[i]
            diffs: list[str] = []
            record: dict = {"对比结果": "有变化"}
            for col in ordered_cols:
                va = df_a.at[idx_a, col] if col in df_a.columns else ""
                vb = df_b.at[idx_b, col] if col in df_b.columns else ""
                ta = _cell_text(va)
                tb = _cell_text(vb)
                record[f"{col}_A"] = va if col in df_a.columns else ""
                record[f"{col}_B"] = vb if col in df_b.columns else ""
                if ta != tb:
                    diffs.append(col)
            if diffs:
                record["变化字段"] = "、".join(diffs)
                changed_rows.append(record)
            else:
                same_count += 1
        for idx in idxs_a[pair_count:]:
            row = {col: df_a.at[idx, col] if col in df_a.columns else "" for col in ordered_cols}
            row["对比结果"] = "仅在表A（同主键多余行）"
            only_a_rows.append(row)
        for idx in idxs_b[pair_count:]:
            row = {col: df_b.at[idx, col] if col in df_b.columns else "" for col in ordered_cols}
            row["对比结果"] = "仅在表B（同主键多余行）"
            only_b_rows.append(row)

    only_a_df = pd.DataFrame(only_a_rows) if only_a_rows else pd.DataFrame(columns=["对比结果", *ordered_cols])
    only_b_df = pd.DataFrame(only_b_rows) if only_b_rows else pd.DataFrame(columns=["对比结果", *ordered_cols])
    changed_df = pd.DataFrame(changed_rows) if changed_rows else pd.DataFrame(columns=["对比结果", "变化字段"])

    summary = pd.DataFrame(
        [
            {"项目": "表A", "内容": path_a.name},
            {"项目": "表B", "内容": path_b.name},
            {"项目": "主键", "内容": " + ".join(key_columns)},
            {"项目": "表A行数", "内容": len(df_a)},
            {"项目": "表B行数", "内容": len(df_b)},
            {"项目": "仅在表A", "内容": len(only_a_rows)},
            {"项目": "仅在表B", "内容": len(only_b_rows)},
            {"项目": "有字段变化", "内容": len(changed_rows)},
            {"项目": "完全相同", "内容": same_count},
            {"项目": "怎么看", "内容": "先看摘要 → 仅在A/仅在B → 有变化（黄标字段）"},
        ]
    )

    combined_parts = [only_a_df, only_b_df, changed_df]
    combined = pd.concat([p for p in combined_parts if not p.empty], ignore_index=True)
    if combined.empty:
        combined = pd.DataFrame(columns=["对比结果", "变化字段", *ordered_cols])

    stem = f"{path_a.stem}_vs_{path_b.stem}"
    result_path, report_path = compare_paths(stem)
    if output_name:
        result_path = ensure_output_dir() / output_name
        report_path = ensure_output_dir() / f"{Path(output_name).stem}_报告.xlsx"

    if progress_cb:
        progress_cb(88, "正在保存对比结果")
    write_excel(combined, result_path, sheet_name="差异清单")

    if progress_cb:
        progress_cb(92, "正在生成对比报告")
    _write_compare_report(report_path, summary, only_a_df, only_b_df, changed_df)
    if progress_cb:
        progress_cb(100, "对比完成！")

    detail = (
        f"仅在A {len(only_a_rows)} 行，仅在B {len(only_b_rows)} 行，"
        f"有变化 {len(changed_rows)} 行，完全相同 {same_count} 行"
    )
    return CompareResult(
        output_path=result_path,
        report_path=report_path,
        only_a=len(only_a_rows),
        only_b=len(only_b_rows),
        changed=len(changed_rows),
        same=same_count,
        detail_text=detail,
    )


def _write_compare_report(
    path: Path,
    summary: pd.DataFrame,
    only_a: pd.DataFrame,
    only_b: pd.DataFrame,
    changed: pd.DataFrame,
) -> None:
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "对比摘要"
    fill_worksheet(ws0, summary)

    ws1 = wb.create_sheet("仅在表A")
    fill_worksheet(ws1, only_a)
    for row_idx in range(2, ws1.max_row + 1):
        for col_idx in range(1, ws1.max_column + 1):
            ws1.cell(row=row_idx, column=col_idx).fill = ORANGE

    ws2 = wb.create_sheet("仅在表B")
    fill_worksheet(ws2, only_b)
    for row_idx in range(2, ws2.max_row + 1):
        for col_idx in range(1, ws2.max_column + 1):
            ws2.cell(row=row_idx, column=col_idx).fill = GREEN

    ws3 = wb.create_sheet("有变化")
    fill_worksheet(ws3, changed)
    # 标黄：同字段 A/B 值不同的单元格
    if not changed.empty:
        headers = [cell.value for cell in ws3[1]]
        for row_idx in range(2, ws3.max_row + 1):
            for col_idx, header in enumerate(headers, start=1):
                if not header or not str(header).endswith("_A"):
                    continue
                base = str(header)[:-2]
                b_header = f"{base}_B"
                if b_header not in headers:
                    continue
                b_idx = headers.index(b_header) + 1
                a_val = _cell_text(ws3.cell(row=row_idx, column=col_idx).value)
                b_val = _cell_text(ws3.cell(row=row_idx, column=b_idx).value)
                if a_val != b_val:
                    ws3.cell(row=row_idx, column=col_idx).fill = YELLOW
                    ws3.cell(row=row_idx, column=b_idx).fill = YELLOW

    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(path)
    except PermissionError as exc:
        raise AppError(
            "无法保存对比报告",
            "请确认报告文件没有被 Excel 打开，并且您有权限写入这个文件夹。",
        ) from exc
