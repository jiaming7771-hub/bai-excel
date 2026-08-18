from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from app_config import ensure_output_dir
from utils.excel_utils import fill_worksheet, read_excel
from utils.error_handler import AppError

PHONE_RE = re.compile(r"^1[3-9]\d{9}$")
EMAIL_RE = re.compile(r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$")
DATE_FORMATS = (
    "%Y-%m-%d",
    "%Y/%m/%d",
    "%Y.%m.%d",
    "%Y%m%d",
    "%Y-%m-%d %H:%M:%S",
    "%Y/%m/%d %H:%M:%S",
    "%Y年%m月%d日",
    "%m/%d/%Y",
    "%d/%m/%Y",
)
# 与 Excel/WPS「填充颜色」油漆桶里的标准黄/橙一致（ARGB）
YELLOW = PatternFill(fill_type="solid", fgColor="FFFFFF00")
ORANGE = PatternFill(fill_type="solid", fgColor="FFFF9800")


@dataclass
class CleanResult:
    output_path: Path
    original_rows: int
    cleaned_rows: int
    deleted_rows: int
    blank_deleted: int = 0
    duplicate_deleted: int = 0
    field_deleted: int = 0
    fixed_cells: int = 0
    review_rows: int = 0
    quality_score: int = 100
    detail_text: str = ""
    quality_summary: str = ""


def _is_blank_value(value) -> bool:
    try:
        if pd.isna(value):
            return True
    except (TypeError, ValueError):
        pass
    text = str(value).strip().lower()
    return text in {"", "nan", "none", "nat", "<na>"}


def _is_blank_row(row: pd.Series) -> bool:
    return all(_is_blank_value(value) for value in row.tolist())


def _text(value) -> str:
    if _is_blank_value(value):
        return ""
    return str(value).strip()


def _digits(value) -> str:
    return re.sub(r"\D", "", _text(value))


def looks_like_phone(value) -> bool:
    digits = _digits(value)
    return len(digits) >= 10 and digits.startswith("1")


def looks_like_email(value) -> bool:
    text = _text(value)
    return "@" in text and "." in text


def looks_like_chinese_name(value) -> bool:
    text = _text(value)
    if not text or looks_like_phone(text) or looks_like_email(text):
        return False
    return bool(re.fullmatch(r"[\u4e00-\u9fff·]{2,8}", text.replace(" ", "")))


def find_column(columns: list[str], hints: tuple[str, ...]) -> str | None:
    for hint in hints:
        for col in columns:
            if hint in str(col):
                return col
    return None


def parse_date_value(value) -> date | None:
    if _is_blank_value(value):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, pd.Timestamp):
        if pd.isna(value):
            return None
        return value.date()

    raw = _text(value)
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            pass

    text = raw.replace("年", "-").replace("月", "-").replace("日", "")
    text = text.replace(".", "-").replace("/", "-")
    try:
        parts = [p for p in re.split(r"[-/\s]", text) if p]
        if len(parts) >= 3 and parts[0].isdigit() and len(parts[0]) == 4:
            return date(int(parts[0]), int(parts[1]), int(parts[2]))
    except Exception:
        return None
    return None


def try_fix_phone(value) -> tuple[object, bool, str | None]:
    if _is_blank_value(value):
        return value, False, None
    original = _text(value)
    digits = _digits(value)
    if PHONE_RE.fullmatch(digits):
        if digits != original:
            return digits, True, None
        return original, False, None
    return value, False, "手机号不正确，需人工核对"


def try_fix_email(value) -> tuple[object, bool, str | None]:
    if _is_blank_value(value):
        return value, False, None
    original = _text(value)
    normalized = original.lower().replace(" ", "")
    if EMAIL_RE.fullmatch(normalized):
        if normalized != original:
            return normalized, True, None
        return original, False, None
    return value, False, "邮箱不正确，需人工核对"


def try_fix_date(value) -> tuple[object, bool, str | None]:
    if _is_blank_value(value):
        return value, False, None
    parsed = parse_date_value(value)
    if parsed is None:
        return value, False, "日期格式不对，需人工核对"
    new_text = parsed.strftime("%Y-%m-%d")
    old_text = _text(value)
    if isinstance(value, (date, datetime, pd.Timestamp)):
        return new_text, False, None
    if old_text == new_text:
        return new_text, False, None
    return new_text, True, None


def classify_contact_value(value) -> str:
    """把单元格内容归类为 name / phone / email / blank / unknown。"""
    if _is_blank_value(value):
        return "blank"
    if looks_like_phone(value):
        return "phone"
    if looks_like_email(value):
        return "email"
    if looks_like_chinese_name(value):
        return "name"
    return "unknown"


def try_fix_shifted_row(
    name_val,
    phone_val,
    email_val,
    *,
    has_name: bool,
    has_phone: bool,
    has_email: bool,
) -> tuple[object, object, object, bool, str | None]:
    """
    尝试把窜行的姓名/手机/邮箱归位。
    仅在「每种类型至多一个、且没有无法归类的挡路值」时自动修复。
    返回 (new_name, new_phone, new_email, fixed, problem)。
    """
    slots: list[tuple[str, object]] = []
    if has_name:
        slots.append(("name", name_val))
    if has_phone:
        slots.append(("phone", phone_val))
    if has_email:
        slots.append(("email", email_val))

    typed: dict[str, object] = {}
    unknowns: list[tuple[str, object]] = []
    misplaced = False

    for expected, val in slots:
        kind = classify_contact_value(val)
        if kind == "blank":
            continue
        if kind == "unknown":
            unknowns.append((expected, val))
            continue
        if kind != expected:
            misplaced = True
        if kind in typed:
            return name_val, phone_val, email_val, False, "疑似窜行但无法唯一归位，需人工核对"
        typed[kind] = val

    if not misplaced:
        return name_val, phone_val, email_val, False, None

    # 有错位时，未知内容占着「本该被归位类型填入」的格子 → 不安全，交人工
    needed = set(typed.keys())
    for expected, _val in unknowns:
        if expected in needed:
            return name_val, phone_val, email_val, False, "疑似窜行但夹杂无法识别内容，需人工核对"

    def home(kind: str, current, enabled: bool):
        if not enabled:
            return current
        if kind in typed:
            return typed[kind]
        kind_now = classify_contact_value(current)
        if kind_now in {"name", "phone", "email"} and kind_now != kind:
            return ""
        return current

    new_name = home("name", name_val, has_name)
    new_phone = home("phone", phone_val, has_phone)
    new_email = home("email", email_val, has_email)

    changed = (
        (has_name and _text(new_name) != _text(name_val))
        or (has_phone and _text(new_phone) != _text(phone_val))
        or (has_email and _text(new_email) != _text(email_val))
    )
    if not changed:
        return name_val, phone_val, email_val, False, "疑似窜行（字段错位），需人工核对"
    return new_name, new_phone, new_email, True, None


def inspect_quality(df: pd.DataFrame) -> dict:
    """清洗前体检：统计空白、重复、窜行、字段异常等（不改数据）。"""
    cols = list(df.columns)
    phone_col = find_column(cols, ("手机", "电话", "mobile", "phone"))
    email_col = find_column(cols, ("邮箱", "邮件", "email", "mail"))
    name_col = find_column(cols, ("姓名", "客户名", "联系人", "名字"))
    date_col = find_column(cols, ("日期", "下单", "时间", "入职"))

    total = len(df)
    blank_rows = int(df.apply(_is_blank_row, axis=1).sum()) if total else 0
    dup_rows = int(df.duplicated(keep="first").sum()) if total else 0

    shifted = bad_phone = bad_email = bad_date = missing_key = 0
    for _, row in df.iterrows():
        name_val = row[name_col] if name_col else None
        phone_val = row[phone_col] if phone_col else None
        email_val = row[email_col] if email_col else None

        if name_col or phone_col or email_col:
            _n, _p, _e, fixed, problem = try_fix_shifted_row(
                name_val,
                phone_val,
                email_val,
                has_name=bool(name_col),
                has_phone=bool(phone_col),
                has_email=bool(email_col),
            )
            if fixed or (problem and "窜行" in problem):
                shifted += 1

        if phone_col and not _is_blank_value(phone_val):
            kind = classify_contact_value(phone_val)
            if kind == "phone":
                _, fixed, problem = try_fix_phone(phone_val)
                if problem:
                    bad_phone += 1
            elif kind != "blank":
                # 可能是窜行，不重复计入手机异常
                if kind not in {"name", "email"}:
                    bad_phone += 1

        if email_col and not _is_blank_value(email_val):
            kind = classify_contact_value(email_val)
            if kind == "email":
                _, fixed, problem = try_fix_email(email_val)
                if problem:
                    bad_email += 1
            elif kind not in {"name", "phone", "blank"}:
                bad_email += 1

        if date_col and not _is_blank_value(row[date_col]):
            _, fixed, problem = try_fix_date(row[date_col])
            if problem:
                bad_date += 1

        name_blank = bool(name_col) and _is_blank_value(name_val)
        phone_blank = bool(phone_col) and _is_blank_value(phone_val)
        if name_col and phone_col:
            if name_blank or phone_blank:
                missing_key += 1
        elif name_blank or phone_blank:
            missing_key += 1

    # 健康度：问题越多分越低
    score = 100
    if total > 0:
        score -= min(25, round(blank_rows / total * 100))
        score -= min(25, round(dup_rows / total * 80))
        score -= min(20, round(shifted / total * 120))
        score -= min(15, round(bad_phone / total * 100))
        score -= min(10, round(bad_email / total * 80))
        score -= min(10, round(bad_date / total * 80))
        score -= min(15, round(missing_key / total * 100))
    score = int(max(0, min(100, score)))

    if score >= 85:
        level = "良好"
        advice = "问题较少，建议抽查「已自动修复」后即可使用。"
    elif score >= 60:
        level = "一般"
        advice = "建议优先核对待人工核对；窜行归位请对照「已自动修复」抽查。"
    else:
        level = "较差"
        advice = "建议先处理「待人工核对」，再抽查自动修复；关键字段缺失需补全后再导入系统。"

    return {
        "total": total,
        "blank_rows": blank_rows,
        "dup_rows": dup_rows,
        "shifted": shifted,
        "bad_phone": bad_phone,
        "bad_email": bad_email,
        "bad_date": bad_date,
        "missing_key": missing_key,
        "score": score,
        "level": level,
        "advice": advice,
        "phone_col": phone_col,
        "email_col": email_col,
        "name_col": name_col,
        "date_col": date_col,
    }


def _dedupe_by_column_with_removed(df: pd.DataFrame, column: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    seen: set[str] = set()
    keep: list = []
    removed: list = []
    for idx, value in df[column].items():
        if _is_blank_value(value):
            keep.append(idx)
            continue
        key = str(value).strip()
        if key in seen:
            removed.append(idx)
            continue
        seen.add(key)
        keep.append(idx)
    return df.loc[keep], df.loc[removed]


def _take_removed(work: pd.DataFrame, mask: pd.Series, reason: str) -> tuple[pd.DataFrame, pd.DataFrame, int]:
    part = work.loc[mask].copy()
    kept = work.loc[~mask].copy()
    count = len(part)
    if count:
        part.insert(0, "删除原因", reason)
    return kept, part, count


def clean_excel(
    path: Path,
    drop_duplicates: bool = False,
    drop_blank_rows: bool = False,
    dedupe_column: str | None = None,
    check_shifted_rows: bool = False,
    fix_phone: bool = False,
    fix_email: bool = False,
    fix_dates: bool = False,
    check_required_fields: bool = False,
    trim_spaces: bool = True,
    output_name: str = "cleaned_result.xlsx",
    progress_cb=None,
) -> CleanResult:
    options = [
        drop_duplicates,
        drop_blank_rows,
        bool(dedupe_column),
        check_shifted_rows,
        fix_phone,
        fix_email,
        fix_dates,
        check_required_fields,
        trim_spaces,
    ]
    if not any(options):
        raise AppError("请先选择清洗方式", "至少勾选一项清洗规则后再开始。")

    if progress_cb:
        progress_cb(6, "正在读取并体检数据")
    original_df = read_excel(path).copy()
    original = len(original_df)
    quality = inspect_quality(original_df)
    work = original_df.copy()
    removed_parts: list[pd.DataFrame] = []
    work["_row_uid"] = range(len(work))
    fixed_map: dict[int, set[str]] = {}
    review_map: dict[int, list[str]] = {}
    review_cols: dict[int, set[str]] = {}
    fix_log: list[dict] = []

    blank_deleted = duplicate_deleted = field_deleted = 0
    fixed_cells = 0

    phone_col = find_column(list(work.columns), ("手机", "电话", "mobile", "phone"))
    email_col = find_column(list(work.columns), ("邮箱", "邮件", "email", "mail"))
    name_col = find_column(list(work.columns), ("姓名", "客户名", "联系人", "名字"))
    date_col = find_column(list(work.columns), ("日期", "下单", "时间", "入职"))

    def mark_fixed(uid: int, column: str, old=None, new=None, reason: str = "已自动修复") -> None:
        nonlocal fixed_cells
        fixed_map.setdefault(uid, set()).add(column)
        fixed_cells += 1
        fix_log.append(
            {
                "uid": uid,
                "字段": column,
                "原值": "" if _is_blank_value(old) else old,
                "新值": "" if _is_blank_value(new) else new,
                "修复说明": reason,
            }
        )

    def mark_review(uid: int, column: str | None, reason: str) -> None:
        review_map.setdefault(uid, [])
        if reason not in review_map[uid]:
            review_map[uid].append(reason)
        if column:
            review_cols.setdefault(uid, set()).add(column)

    if drop_blank_rows:
        if progress_cb:
            progress_cb(16, "正在删除空白行")
        data_cols = [c for c in work.columns if c != "_row_uid"]
        mask = work[data_cols].apply(_is_blank_row, axis=1)
        work, part, blank_deleted = _take_removed(work, mask, "空白行")
        if blank_deleted:
            part = part.drop(columns=["_row_uid"], errors="ignore")
            removed_parts.append(part)

    if trim_spaces and name_col:
        if progress_cb:
            progress_cb(24, "正在清理姓名空格")
        for idx in work.index:
            value = work.at[idx, name_col]
            if _is_blank_value(value) or not isinstance(value, str):
                continue
            stripped = value.strip()
            if stripped != value:
                uid = int(work.at[idx, "_row_uid"])
                work.at[idx, name_col] = stripped
                mark_fixed(uid, name_col, value, stripped, "姓名去掉首尾空格")

    if check_shifted_rows and (name_col or phone_col or email_col):
        if progress_cb:
            progress_cb(34, "正在检查并尝试归位窜行")
        for idx in work.index:
            uid = int(work.at[idx, "_row_uid"])
            phone_val = work.at[idx, phone_col] if phone_col else None
            email_val = work.at[idx, email_col] if email_col else None
            name_val = work.at[idx, name_col] if name_col else None
            new_name, new_phone, new_email, fixed, problem = try_fix_shifted_row(
                name_val,
                phone_val,
                email_val,
                has_name=bool(name_col),
                has_phone=bool(phone_col),
                has_email=bool(email_col),
            )
            if fixed:
                reason = "窜行自动归位（按内容换回正确列）"
                if name_col and _text(new_name) != _text(name_val):
                    work.at[idx, name_col] = new_name
                    mark_fixed(uid, name_col, name_val, new_name, reason)
                if phone_col and _text(new_phone) != _text(phone_val):
                    work.at[idx, phone_col] = new_phone
                    mark_fixed(uid, phone_col, phone_val, new_phone, reason)
                if email_col and _text(new_email) != _text(email_val):
                    work.at[idx, email_col] = new_email
                    mark_fixed(uid, email_col, email_val, new_email, reason)
            elif problem:
                for col, val, expected in (
                    (name_col, name_val, "name"),
                    (phone_col, phone_val, "phone"),
                    (email_col, email_val, "email"),
                ):
                    if not col or _is_blank_value(val):
                        continue
                    kind = classify_contact_value(val)
                    if kind != expected:
                        mark_review(uid, col, problem)

    if fix_phone and phone_col:
        if progress_cb:
            progress_cb(46, "正在检查手机号")
        for idx in work.index:
            uid = int(work.at[idx, "_row_uid"])
            if any("窜行" in r for r in review_map.get(uid, [])):
                continue
            old_val = work.at[idx, phone_col]
            new_val, fixed, problem = try_fix_phone(old_val)
            if fixed:
                work.at[idx, phone_col] = new_val
                mark_fixed(uid, phone_col, old_val, new_val, "手机号格式整理（去空格/横线等）")
            elif problem:
                mark_review(uid, phone_col, problem)

    if fix_email and email_col:
        if progress_cb:
            progress_cb(56, "正在检查邮箱")
        for idx in work.index:
            uid = int(work.at[idx, "_row_uid"])
            if any("窜行" in r for r in review_map.get(uid, [])):
                continue
            old_val = work.at[idx, email_col]
            new_val, fixed, problem = try_fix_email(old_val)
            if fixed:
                work.at[idx, email_col] = new_val
                mark_fixed(uid, email_col, old_val, new_val, "邮箱格式整理（去空格并转小写）")
            elif problem:
                mark_review(uid, email_col, problem)

    if fix_dates and date_col:
        if progress_cb:
            progress_cb(66, "正在检查日期")
        for idx in work.index:
            uid = int(work.at[idx, "_row_uid"])
            old_val = work.at[idx, date_col]
            new_val, fixed, problem = try_fix_date(old_val)
            if fixed:
                work.at[idx, date_col] = new_val
                mark_fixed(uid, date_col, old_val, new_val, "日期统一为 YYYY-MM-DD")
            elif problem:
                mark_review(uid, date_col, problem)
            elif not _is_blank_value(old_val):
                work.at[idx, date_col] = new_val

    if check_required_fields and (name_col or phone_col):
        if progress_cb:
            progress_cb(72, "正在检查关键字段缺失")
        for idx in work.index:
            uid = int(work.at[idx, "_row_uid"])
            if name_col and _is_blank_value(work.at[idx, name_col]):
                mark_review(uid, name_col, "关键字段缺失：姓名为空，需人工补全")
            if phone_col and _is_blank_value(work.at[idx, phone_col]):
                mark_review(uid, phone_col, "关键字段缺失：手机号为空，需人工补全")

    if drop_duplicates:
        if progress_cb:
            progress_cb(76, "正在删除完全重复行")
        data_cols = [c for c in work.columns if c != "_row_uid"]
        duplicated = work.duplicated(subset=data_cols, keep="first")
        work, part, duplicate_deleted = _take_removed(work, duplicated, "完全重复行")
        if duplicate_deleted:
            part = part.drop(columns=["_row_uid"], errors="ignore")
            removed_parts.append(part)

    if dedupe_column:
        if progress_cb:
            progress_cb(84, f"正在按 {dedupe_column} 去重")
        if dedupe_column not in work.columns:
            raise AppError("找不到去重字段", "请选择表格里实际存在的字段，例如手机号、姓名或订单号。")
        kept, field_part = _dedupe_by_column_with_removed(work, dedupe_column)
        if not field_part.empty:
            field_part = field_part.drop(columns=["_row_uid"], errors="ignore").copy()
            field_part.insert(0, "删除原因", f"按{dedupe_column}去重（已保留第一条）")
            removed_parts.append(field_part)
            field_deleted = len(field_part)
        work = kept

    review_records = []
    for idx in work.index:
        uid = int(work.at[idx, "_row_uid"])
        problems = review_map.get(uid)
        if not problems:
            continue
        row = work.loc[idx].drop(labels=["_row_uid"]).to_dict()
        review_records.append({"问题说明": "；".join(problems), **row})

    uid_list = [int(v) for v in work["_row_uid"].tolist()]
    cleaned_df = work.drop(columns=["_row_uid"]).reset_index(drop=True)
    highlight_fixed: dict[int, set[str]] = {}
    highlight_review: dict[int, set[str]] = {}
    markers: list[str] = []
    uid_to_result_row = {uid: i + 1 for i, uid in enumerate(uid_list)}
    for final_i, uid in enumerate(uid_list):
        tags: list[str] = []
        if uid in fixed_map:
            highlight_fixed[final_i] = fixed_map[uid]
            tags.append("已修复")
        if uid in review_cols:
            highlight_review[final_i] = review_cols[uid]
            tags.append("待核对")
        markers.append("；".join(tags))
    cleaned_df.insert(0, "清洗标记", markers)

    kept_uids = set(uid_list)
    fixed_records = []
    for item in fix_log:
        uid = int(item["uid"])
        if uid not in kept_uids:
            continue
        fixed_records.append(
            {
                "结果行号": uid_to_result_row[uid],
                "字段": item["字段"],
                "原值": item["原值"],
                "新值": item["新值"],
                "修复说明": item["修复说明"],
            }
        )
    fixed_df = (
        pd.DataFrame(fixed_records)
        if fixed_records
        else pd.DataFrame(columns=["结果行号", "字段", "原值", "新值", "修复说明"])
    )
    if not fixed_df.empty:
        fixed_df = fixed_df.sort_values(["结果行号", "字段"], kind="stable").reset_index(drop=True)
    # 与最终仍保留在结果中的修复次数对齐
    fixed_cells = len(fixed_df)

    deleted_df = (
        pd.concat(removed_parts, ignore_index=True)
        if removed_parts
        else pd.DataFrame(columns=["删除原因", *list(original_df.columns)])
    )
    review_df = (
        pd.DataFrame(review_records)
        if review_records
        else pd.DataFrame(columns=["问题说明", *list(cleaned_df.columns)])
    )

    cleaned = len(cleaned_df)
    deleted = max(0, original - cleaned)
    review_count = len(review_df)
    score = int(quality["score"])
    level = quality["level"]
    detail = (
        f"数据健康度 {score} 分（{level}）；"
        f"删除 {deleted} 行；已修复 {fixed_cells} 处；待人工核对 {review_count} 行"
    )
    quality_summary = (
        f"体检：空白{quality['blank_rows']} / 重复{quality['dup_rows']} / "
        f"窜行{quality['shifted']} / 手机异常{quality['bad_phone']} / "
        f"邮箱异常{quality['bad_email']} / 日期异常{quality['bad_date']} / "
        f"关键缺失{quality['missing_key']}。{quality['advice']}"
    )

    quality_df = pd.DataFrame(
        [
            {"类别": "清洗前体检", "项目": "总行数", "数量": quality["total"], "说明": "原始数据行数"},
            {"类别": "清洗前体检", "项目": "空白行", "数量": quality["blank_rows"], "说明": "整行无内容，将删除"},
            {"类别": "清洗前体检", "项目": "完全重复行", "数量": quality["dup_rows"], "说明": "与前面完全相同，将删除（保留首条）"},
            {"类别": "清洗前体检", "项目": "疑似窜行", "数量": quality["shifted"], "说明": "姓名/手机/邮箱疑似错位，能修则归位"},
            {"类别": "清洗前体检", "项目": "手机号异常", "数量": quality["bad_phone"], "说明": "格式不对或无法识别"},
            {"类别": "清洗前体检", "项目": "邮箱异常", "数量": quality["bad_email"], "说明": "格式不对或无法识别"},
            {"类别": "清洗前体检", "项目": "日期异常", "数量": quality["bad_date"], "说明": "无法解析为标准日期"},
            {"类别": "清洗前体检", "项目": "关键字段缺失", "数量": quality["missing_key"], "说明": "姓名或手机号为空"},
            {"类别": "清洗前体检", "项目": "数据健康度", "数量": score, "说明": f"{level}（100分制）"},
            {"类别": "本次处理", "项目": "清洗后行数", "数量": cleaned, "说明": "结果表可用行数"},
            {"类别": "本次处理", "项目": "已删除", "数量": deleted, "说明": "见「已删除记录」"},
            {"类别": "本次处理", "项目": "已自动修复", "数量": fixed_cells, "说明": "见「已自动修复」原值→新值"},
            {"类别": "本次处理", "项目": "待人工核对", "数量": review_count, "说明": "见「待人工核对」"},
            {"类别": "建议", "项目": "下一步", "数量": "-", "说明": quality["advice"]},
        ]
    )

    output = ensure_output_dir() / output_name
    if progress_cb:
        progress_cb(92, "正在生成清洗报告")
    _write_clean_report(
        output,
        cleaned_df=cleaned_df,
        quality_df=quality_df,
        fixed_df=fixed_df,
        deleted_df=deleted_df,
        review_df=review_df,
        highlight_fixed=highlight_fixed,
        highlight_review=highlight_review,
        original_rows=original,
        cleaned_rows=cleaned,
        deleted_rows=deleted,
        blank_deleted=blank_deleted,
        duplicate_deleted=duplicate_deleted,
        field_deleted=field_deleted,
        fixed_cells=fixed_cells,
        review_rows=review_count,
        quality_score=score,
        dedupe_column=dedupe_column,
        source_name=Path(path).name,
    )
    if progress_cb:
        progress_cb(100, "清洗完成！")
    return CleanResult(
        output_path=output,
        original_rows=original,
        cleaned_rows=cleaned,
        deleted_rows=deleted,
        blank_deleted=blank_deleted,
        duplicate_deleted=duplicate_deleted,
        field_deleted=field_deleted,
        fixed_cells=fixed_cells,
        review_rows=review_count,
        quality_score=score,
        detail_text=detail,
        quality_summary=quality_summary,
    )


def _write_clean_report(
    path: Path,
    *,
    cleaned_df: pd.DataFrame,
    quality_df: pd.DataFrame,
    fixed_df: pd.DataFrame,
    deleted_df: pd.DataFrame,
    review_df: pd.DataFrame,
    highlight_fixed: dict[int, set[str]],
    highlight_review: dict[int, set[str]],
    original_rows: int,
    cleaned_rows: int,
    deleted_rows: int,
    blank_deleted: int,
    duplicate_deleted: int,
    field_deleted: int,
    fixed_cells: int,
    review_rows: int,
    quality_score: int,
    dedupe_column: str | None,
    source_name: str,
) -> None:
    wb = Workbook()

    ws_q = wb.active
    ws_q.title = "清洗报告"
    fill_worksheet(ws_q, quality_df)
    # 健康度那一行强调颜色
    for row_idx in range(2, ws_q.max_row + 1):
        item = ws_q.cell(row=row_idx, column=2).value
        if item == "数据健康度":
            fill = YELLOW if quality_score >= 60 else ORANGE
            for col_idx in range(1, ws_q.max_column + 1):
                ws_q.cell(row=row_idx, column=col_idx).fill = fill

    ws1 = wb.create_sheet("清洗结果")
    fill_worksheet(ws1, cleaned_df)
    _apply_row_fills(ws1, cleaned_df)
    _apply_highlights(ws1, cleaned_df, highlight_fixed, YELLOW)
    _apply_highlights(ws1, cleaned_df, highlight_review, ORANGE)

    ws_fixed = wb.create_sheet("已自动修复")
    fill_worksheet(ws_fixed, fixed_df)
    if not fixed_df.empty:
        for row_idx in range(2, ws_fixed.max_row + 1):
            for col_idx in range(1, ws_fixed.max_column + 1):
                ws_fixed.cell(row=row_idx, column=col_idx).fill = YELLOW

    ws2 = wb.create_sheet("待人工核对")
    fill_worksheet(ws2, review_df)
    if not review_df.empty:
        for row_idx in range(2, ws2.max_row + 1):
            ws2.cell(row=row_idx, column=1).fill = ORANGE

    ws3 = wb.create_sheet("已删除记录")
    fill_worksheet(ws3, deleted_df)

    summary = pd.DataFrame(
        [
            {"项目": "原文件", "内容": source_name},
            {"项目": "原始行数", "内容": original_rows},
            {"项目": "清洗后行数", "内容": cleaned_rows},
            {"项目": "删除合计", "内容": deleted_rows},
            {"项目": "空白行（已删除）", "内容": blank_deleted},
            {"项目": "完全重复行（已删除）", "内容": duplicate_deleted},
            {"项目": "字段去重（已删除）", "内容": field_deleted},
            {"项目": "已修复并标黄", "内容": fixed_cells},
            {"项目": "待人工核对行数", "内容": review_rows},
            {"项目": "数据健康度", "内容": quality_score},
            {"项目": "去重字段", "内容": dedupe_column or "未使用"},
            {"项目": "怎么核对", "内容": "先看「清洗报告」→「已自动修复」→「待人工核对」→「清洗结果」"},
        ]
    )
    ws4 = wb.create_sheet("清洗说明")
    fill_worksheet(ws4, summary)

    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(path)
    except PermissionError as exc:
        raise AppError(
            "无法保存结果文件",
            "请确认输出文件没有被 Excel 打开，并且您有权限写入这个文件夹。",
        ) from exc


def _apply_row_fills(ws, df: pd.DataFrame) -> None:
    """按「清洗标记」给整行填充底色（等同于 WPS/Excel 油漆桶填充）。"""
    if df.empty or "清洗标记" not in df.columns:
        return
    ncols = len(df.columns)
    for row_pos, marker in enumerate(df["清洗标记"].tolist()):
        text = str(marker or "")
        if not text:
            continue
        fill = ORANGE if "待核对" in text else YELLOW if "已修复" in text else None
        if fill is None:
            continue
        excel_row = row_pos + 2
        for col in range(1, ncols + 1):
            ws.cell(row=excel_row, column=col).fill = fill


def _apply_highlights(ws, df: pd.DataFrame, mapping: dict[int, set[str]], fill: PatternFill) -> None:
    if df.empty or not mapping:
        return
    col_index = {name: i for i, name in enumerate(df.columns, start=1)}
    for row_pos, columns in mapping.items():
        excel_row = row_pos + 2
        for column in columns:
            idx = col_index.get(column)
            if idx:
                ws.cell(row=excel_row, column=idx).fill = fill
